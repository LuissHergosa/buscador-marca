"""
Excel export service for brand detection results.
"""

from datetime import datetime
from io import BytesIO
from typing import List
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

from app.models.document import Document
from app.models.brand_detection import BrandDetection


class ExcelService:
    """Service for generating Excel reports from brand detection results."""

    def __init__(self):
        self.wb = None
        self.ws = None

    def generate_document_results_excel(self, document: Document) -> BytesIO:
        """
        Generate an Excel file with all brand detection results for a document.

        Args:
            document: Document with results to export

        Returns:
            BytesIO: Excel file content as bytes
        """
        # Create workbook and worksheet
        self.wb = openpyxl.Workbook()
        self.ws = self.wb.active
        self.ws.title = "Análisis de Marcas"

        # Set up styles
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(
            start_color="4472C4", end_color="4472C4", fill_type="solid"
        )
        border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        # Document information section
        self._add_document_info(document, header_font, header_fill, border)

        # Results section
        self._add_results_table(document, header_font, header_fill, border)

        # Statistics section
        self._add_statistics_section(document, header_font, header_fill, border)

        # Adjust column widths
        self._adjust_column_widths()

        # Save to BytesIO
        excel_buffer = BytesIO()
        self.wb.save(excel_buffer)
        excel_buffer.seek(0)

        return excel_buffer

    def _add_document_info(
        self,
        document: Document,
        header_font: Font,
        header_fill: PatternFill,
        border: Border,
    ):
        """Add document information section."""
        # Document info header
        self.ws["A1"] = "INFORMACIÓN DEL DOCUMENTO"
        self.ws["A1"].font = header_font
        self.ws["A1"].fill = header_fill
        self.ws["A1"].border = border
        self.ws.merge_cells("A1:B1")

        # Document details
        info_data = [
            ("Nombre del archivo:", document.filename),
            ("Fecha de subida:", document.upload_date.strftime("%d/%m/%Y %H:%M")),
            ("Total de páginas:", str(document.total_pages)),
            ("Estado:", document.status),
            ("Fecha de exportación:", datetime.now().strftime("%d/%m/%Y %H:%M")),
        ]

        for i, (label, value) in enumerate(info_data, start=2):
            self.ws[f"A{i}"] = label
            self.ws[f"B{i}"] = value
            self.ws[f"A{i}"].font = Font(bold=True)
            self.ws[f"A{i}"].border = border
            self.ws[f"B{i}"].border = border

    def _add_results_table(
        self,
        document: Document,
        header_font: Font,
        header_fill: PatternFill,
        border: Border,
    ):
        """Add the main results table."""
        # Start results table after document info (row 8)
        start_row = 8

        # Results header
        self.ws[f"A{start_row}"] = "RESULTADOS DEL ANÁLISIS"
        self.ws[f"A{start_row}"].font = header_font
        self.ws[f"A{start_row}"].fill = header_fill
        self.ws[f"A{start_row}"].border = border
        self.ws.merge_cells(f"A{start_row}:C{start_row}")

        # Table headers
        headers = ["Página", "Marca Detectada", "Estado de Revisión"]
        header_row = start_row + 2

        for col, header in enumerate(headers, start=1):
            cell = self.ws.cell(row=header_row, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
            cell.alignment = Alignment(horizontal="center")

        # Add results data
        current_row = header_row + 1

        if document.results:
            # Sort results by page number
            sorted_results = sorted(document.results, key=lambda x: x.page_number)

            for result in sorted_results:
                if result.brands_detected:
                    # Sort brands alphabetically within each page
                    sorted_brands = sorted(result.brands_detected)

                    for brand in sorted_brands:
                        # Get review status
                        is_reviewed = result.brands_review_status.get(brand, False)
                        review_status = "Revisado" if is_reviewed else "Por Revisar"

                        # Add row data
                        row_data = [result.page_number, brand, review_status]

                        for col, value in enumerate(row_data, start=1):
                            cell = self.ws.cell(
                                row=current_row, column=col, value=value
                            )
                            cell.border = border
                            if col == 3:  # Review status column
                                if is_reviewed:
                                    cell.fill = PatternFill(
                                        start_color="C6EFCE",
                                        end_color="C6EFCE",
                                        fill_type="solid",
                                    )
                                else:
                                    cell.fill = PatternFill(
                                        start_color="FFC7CE",
                                        end_color="FFC7CE",
                                        fill_type="solid",
                                    )

                        current_row += 1

    def _add_statistics_section(
        self,
        document: Document,
        header_font: Font,
        header_fill: PatternFill,
        border: Border,
    ):
        """Add statistics section."""
        # Calculate statistics
        total_brands = 0
        reviewed_brands = 0
        pages_with_results = 0

        if document.results:
            for result in document.results:
                if result.brands_detected:
                    pages_with_results += 1
                    total_brands += len(result.brands_detected)
                    reviewed_brands += sum(
                        1
                        for brand in result.brands_detected
                        if result.brands_review_status.get(brand, False)
                    )

        # Find next available row
        start_row = self.ws.max_row + 3

        # Statistics header
        self.ws[f"A{start_row}"] = "ESTADÍSTICAS"
        self.ws[f"A{start_row}"].font = header_font
        self.ws[f"A{start_row}"].fill = header_fill
        self.ws[f"A{start_row}"].border = border
        self.ws.merge_cells(f"A{start_row}:B{start_row}")

        # Statistics data
        stats_data = [
            ("Total de marcas detectadas:", str(total_brands)),
            ("Marcas revisadas:", str(reviewed_brands)),
            ("Marcas por revisar:", str(total_brands - reviewed_brands)),
            (
                "Páginas con resultados:",
                f"{pages_with_results} de {document.total_pages}",
            ),
            (
                "Porcentaje de revisión:",
                f"{(reviewed_brands / total_brands * 100):.1f}%"
                if total_brands > 0
                else "0%",
            ),
        ]

        for i, (label, value) in enumerate(stats_data, start=start_row + 2):
            self.ws[f"A{i}"] = label
            self.ws[f"B{i}"] = value
            self.ws[f"A{i}"].font = Font(bold=True)
            self.ws[f"A{i}"].border = border
            self.ws[f"B{i}"].border = border

    def _adjust_column_widths(self):
        """Adjust column widths for better readability."""
        # Auto-adjust column widths based on content
        for column in self.ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)

            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass

            # Set minimum and maximum widths
            adjusted_width = min(max(max_length + 2, 12), 50)
            self.ws.column_dimensions[column_letter].width = adjusted_width

    def generate_filename(self, document: Document) -> str:
        """Generate a descriptive filename for the Excel export."""
        # Clean filename (remove extension and special characters)
        clean_name = document.filename.replace(".pdf", "").replace(" ", "_")
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        return f"{clean_name}_resultados_{timestamp}.xlsx"
